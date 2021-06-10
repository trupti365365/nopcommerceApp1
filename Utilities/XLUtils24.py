import openpyxl

def getRowCount(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(SheetName)
    return(sheet.max_row)

def getColumnCount(file,SheetName):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(SheetName)
    return(sheet.max_column)
def readData(file,SheetName,rowno,columnno):
    workbook=openpyxl.load_workbook(file)
    sheet=workbook.get_sheet_by_name(SheetName)
    return sheet.cell(row=rowno,column=columnno).value

def writeData(file,SheetName,rowno,columnno,data):
    workbook = openpyxl.load_workbook(file)
    sheet = workbook.get_sheet_by_name(SheetName)
    return sheet.cell(row=rowno, column=columnno).value
    sheet.cell(row=rowno, column=columnno).value = "data"
    workbook.save(file)



